import os
import re
import openpyxl as xl



# Convert .csv file to .xlsx file
print('make sure the file to convert is in the NoBadWord folder')
print("plz enter the file's name you want to convert")
while True:
    userFile = input('>')
    HaveFile = True
    try:
        wb = xl.load_workbook(f'{userFile}.xlsx')
    except FileNotFoundError:
        print('File not found')
        HaveFile = False
    if HaveFile:
        break

changeRegex = re.compile(r';;;;;;')
sheet = wb['Sheet1']
for roll in range(1, sheet.max_row + 1):
    cell = sheet.cell(roll, 1)
    fix_cell_value = changeRegex.sub(r'', f"{cell.value}")
    print(fix_cell_value)
    cell.value = fix_cell_value
wb.save(f'{userFile}).xlsx')