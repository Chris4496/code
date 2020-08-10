import re
import openpyxl as xl


def curseWordRegex(UserChat):
    for roll in range(1, sheet.max_row + 1):
        cell = sheet.cell(roll, 1)
        bad_regex = re.compile(f'{cell.value}')
        mo = bad_regex.search(UserChat)
        if mo == None:
            pass
        else:
            return True


def replaceCurseWords(UserChat):
    for roll in range(1, sheet.max_row + 1):
        cell = sheet.cell(roll, 1)
        badRegex = re.compile(fr'{cell.value}')
        WordLen = len(cell.value)
        UserChat = badRegex.sub('#' * WordLen, UserChat)
    return UserChat


while True:
    word_fliter = input('(C)ustom or (G)eneral Bad Words? ') # Use General for better result
    if word_fliter.lower() == 'c':
        wb = xl.load_workbook('badwordss.xlsx')
        break
    if word_fliter.lower() == 'g':
        wb = xl.load_workbook('full-list-of-bad-words_csv-file_2018_07_30.xlsx')
        break
    else:
        print('unknown input')
sheet = wb['Sheet1']
while True:
    UserChat = input('>')
    FixedUserChat = UserChat.lower()
    HaveFuck = curseWordRegex(FixedUserChat)
    if HaveFuck:
        print(replaceCurseWords(FixedUserChat))
    if not HaveFuck:
        print(UserChat)